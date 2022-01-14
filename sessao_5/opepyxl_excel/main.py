import openpyxl as excel
import pandas as pd
from random import uniform

pedido = excel.load_workbook('pedidos.xlsx')
planilhas = pedido.sheetnames
# print(planilhas)

planilhaas = pedido['Página1']
# print(planilha1['b4'].value)

tabela = pd.read_excel('pedidos.xlsx')

"""
#acessa todos os valores da coluna 'b'
for campo in planilha1['b']:
  print(campo.value)
"""
"""
#acessa range de a1 até c2, por exemplo

for linha in planilha1['a1:c2']:
  for coluna in linha:
    print(coluna.value)
"""
"""
for linha in planilha1:
  if linha[0].value is not None:
    print(linha[0].value, end=' ')

  if linha[1].value is not None:
    print(linha[1].value, end=' ')

  if linha[2].value is not None:
    print(linha[2].value)

"""
"""
#alterar valores da planilha sem alterar o valor da planilha original

print(planilha1['B3'].value)

planilha1['B3'].value = 2200  

print(planilha1['B3'].value)

print(tabela)
tabela2 = pd.read_excel('nova_planilha.xlsx')

print("-"*65)

aux = 3
for linha in range(5, 16):

  planilha1.cell(linha, 1).value = aux + 1
  planilha1.cell(linha, 2).value = 1200 + linha

  preco = round(uniform(10, 100), 2)
  planilha1.cell(linha, 3).value = preco
pedido.save('nova_planilha.xlsx')
print(tabela2)

"""

planilha = excel.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

aux = 0
for linha in range(1, 11):
    numero_peido = linha - 1

    planilha1.cell(linha, 1).value = numero_peido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco
for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Diego {linha} round(uniform(10,100)2)'
    planilha2.cell(linha, 2).value = f'José {linha} round(uniform(10,100)2)'

planilha.save("nova_planilha2.xlsx")

tabela2 = pd.read_excel("nova_planilha2.xlsx")
print(tabela2)

tabela3 = pd.read_excel('nova_planilha2.xlsx', sheet_name='Planilha2')

print('Planilha 2')
print(tabela3[::])
